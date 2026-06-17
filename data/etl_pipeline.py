"""
Enterprise Financial Portfolio — ETL Pipeline
==============================================
Extracts data from public sources, transforms it, and loads it into
the EnterpriseFinancialPortfolio_Data.xlsx data file.

Data Sources:
  - SEC EDGAR: https://efts.sec.gov/LATEST/search-index
  - FRED:      https://fred.stlouisfed.org/series/
  - NSE:       https://www.nse.co.ke/market-statistics
  - World Bank: https://api.worldbank.org/v2/

Author  : Phil Kibet | Biostatistician & Financial Data Analyst
Updated : June 2026

Usage:
    pip install requests pandas openpyxl
    python etl_pipeline.py
"""

import os
import json
import random
import math
from datetime import date, timedelta

# Safe imports with validation
try:
    import pandas as pd
    print("[OK] pandas loaded")
except ImportError:
    raise ImportError("Run: pip install pandas")

try:
    import openpyxl
    print("[OK] openpyxl loaded")
except ImportError:
    raise ImportError("Run: pip install openpyxl")

# Optional: requests for live API pulls
try:
    import requests
    REQUESTS_AVAILABLE = True
    print("[OK] requests loaded — live API pulls enabled")
except ImportError:
    REQUESTS_AVAILABLE = False
    print("[WARN] requests not available — using synthetic data")


# ── CONFIGURATION ─────────────────────────────────────────────────────────────
OUTPUT_PATH = "EnterpriseFinancialPortfolio_Data.xlsx"

COMPANIES = [
    (1,  "AAPL",  "Apple Inc.",          "AAPL",  "Technology",         "Tech Hardware",     "Americas", "USA",         "NASDAQ", "3571", 147000,  15441943070),
    (2,  "MSFT",  "Microsoft Corp.",      "MSFT",  "Technology",         "Software",          "Americas", "USA",         "NASDAQ", "7372", 221000,  7432000000),
    (3,  "GOOGL", "Alphabet Inc.",        "GOOGL", "Technology",         "Internet Services", "Americas", "USA",         "NASDAQ", "7375", 182502,  5500000000),
    (4,  "JPM",   "JPMorgan Chase",       "JPM",   "Financials",         "Diversified Banks", "Americas", "USA",         "NYSE",   "6020", 295793,  2900000000),
    (5,  "XOM",   "ExxonMobil",           "XOM",   "Energy",             "Oil & Gas",         "Americas", "USA",         "NYSE",   "2911",  62000,  4270000000),
    (6,  "JNJ",   "Johnson & Johnson",    "JNJ",   "Healthcare",         "Pharmaceuticals",   "Americas", "USA",         "NYSE",   "2836", 152700,  2683731854),
    (7,  "SCOM",  "Safaricom PLC",        "SCOM",  "Comm. Services",     "Telecom",           "EMEA",     "Kenya",       "NSE",    "4813",   6000, 40065428272),
    (8,  "EQTY",  "Equity Group",         "EQTY",  "Financials",         "Commercial Banks",  "EMEA",     "Kenya",       "NSE",    "6022",   9000,  3771905600),
    (9,  "UBSN",  "UBS Group",            "UBSN",  "Financials",         "Investment Banks",  "EMEA",     "Switzerland", "LSE",    "6211", 111543,  3859015385),
    (10, "SHEL",  "Shell PLC",            "SHEL",  "Energy",             "Oil & Gas",         "EMEA",     "UK",          "LSE",    "2911",  93000,  6664338036),
]

TICKERS = [
    (1,  "AAPL",  "Apple Inc.",       "Technology",   "Tech Hardware",     "NASDAQ", "USD", "Mega Cap"),
    (2,  "MSFT",  "Microsoft Corp.",  "Technology",   "Software",          "NASDAQ", "USD", "Mega Cap"),
    (3,  "GOOGL", "Alphabet Inc.",    "Technology",   "Internet Services", "NASDAQ", "USD", "Mega Cap"),
    (4,  "JPM",   "JPMorgan Chase",   "Financials",   "Diversified Banks", "NYSE",   "USD", "Mega Cap"),
    (5,  "XOM",   "ExxonMobil",       "Energy",       "Oil & Gas",         "NYSE",   "USD", "Mega Cap"),
    (6,  "JNJ",   "J&J",              "Healthcare",   "Pharmaceuticals",   "NYSE",   "USD", "Mega Cap"),
    (7,  "SCOM",  "Safaricom PLC",    "Comm. Svc",    "Telecom",           "NSE",    "KES", "Large Cap"),
    (8,  "EQTY",  "Equity Group",     "Financials",   "Comm. Banks",       "NSE",    "KES", "Large Cap"),
    (9,  "UBSN",  "UBS Group",        "Financials",   "Inv. Banks",        "LSE",    "CHF", "Mega Cap"),
    (10, "SHEL",  "Shell PLC",        "Energy",       "Oil & Gas",         "LSE",    "USD", "Mega Cap"),
]

DEPARTMENTS = [
    (1, "Sales",      "REV-001", "Commercial", "Americas", "James Smith"),
    (2, "Marketing",  "MKT-001", "Commercial", "Americas", "Aisha Johnson"),
    (3, "Operations", "OPS-001", "Operations", "Global",   "Brian Chen"),
    (4, "R&D",        "RD-001",  "Innovation", "EMEA",     "Chioma Osei"),
    (5, "Finance",    "FIN-001", "Corporate",  "Americas", "Dev Patel"),
    (6, "HR",         "HR-001",  "Corporate",  "Global",   "Eva Muller"),
    (7, "IT",         "IT-001",  "Technology", "Global",   "Frank Kim"),
    (8, "Legal",      "LEG-001", "Corporate",  "Americas", "Grace Williams"),
]

CURRENCIES = ["USD", "EUR", "GBP", "JPY", "KES", "AUD", "CHF"]
COUNTRIES   = [("USA", 3.1, 3.4, 5.25, 3.7, 104.2, 16.8),
               ("UK",  0.4, 4.6, 5.25, 4.2, None,  None),
               ("EU",  0.5, 2.9, 4.50, 6.4, None,  None),
               ("Kenya",5.6,6.8,12.50, 5.5, None,  None),
               ("Switzerland",1.2,1.7,1.75,2.1,None,None)]


def rnd(val, dp=2):
    return round(val, dp)


def gauss(mu=0, sigma=1):
    return random.gauss(mu, sigma)


# ── DimDate ───────────────────────────────────────────────────────────────────
def build_dim_date():
    print("Building DimDate...")
    rows = []
    months = ["January","February","March","April","May","June",
              "July","August","September","October","November","December"]
    quarters = {1:"Q1",2:"Q1",3:"Q1",4:"Q2",5:"Q2",6:"Q2",
                7:"Q3",8:"Q3",9:"Q3",10:"Q4",11:"Q4",12:"Q4"}
    fq = {m: "FQ"+q[1] for m,q in quarters.items()}
    d = date(2019, 1, 1)
    while d <= date(2024, 12, 31):
        rows.append({
            "date_key":     int(d.strftime("%Y%m%d")),
            "Date":         d,
            "Year":         d.year,
            "Quarter":      quarters[d.month],
            "Month":        d.month,
            "MonthName":    months[d.month-1],
            "MonthShort":   months[d.month-1][:3],
            "FiscalYear":   d.year,
            "FiscalQuarter":fq[d.month],
            "WeekNumber":   d.isocalendar()[1],
            "DayOfWeek":    d.strftime("%A"),
            "IsWeekend":    d.weekday() >= 5,
            "IsTradingDay": d.weekday() < 5,
            "YearMonth":    f"{d.year}-{months[d.month-1][:3]}"
        })
        d += timedelta(days=1)
    print(f"  DimDate: {len(rows):,} rows")
    return pd.DataFrame(rows)


# ── DimCompany ────────────────────────────────────────────────────────────────
def build_dim_company():
    print("Building DimCompany...")
    rows = []
    for co in COMPANIES:
        (key, co_id, name, ticker, sector, industry,
         region, country, exchange, sic, emp, shares) = co
        rows.append({
            "company_key": key, "company_id": co_id,
            "company_name": name, "ticker": ticker,
            "sector": sector, "industry": industry,
            "region": region, "country": country,
            "exchange": exchange, "sic_code": sic,
            "is_current": True,
            "employee_count": emp, "shares_outstanding": shares,
            "effective_from": date(2019,1,1),
            "effective_to":   date(9999,12,31)
        })
    print(f"  DimCompany: {len(rows)} rows")
    return pd.DataFrame(rows)


# ── DimTicker ─────────────────────────────────────────────────────────────────
def build_dim_ticker():
    print("Building DimTicker...")
    rows = [{"ticker_key":t[0],"ticker":t[1],"company_name":t[2],
              "gics_sector":t[3],"gics_industry":t[4],"exchange":t[5],
              "currency":t[6],"market_cap_band":t[7]} for t in TICKERS]
    print(f"  DimTicker: {len(rows)} rows")
    return pd.DataFrame(rows)


# ── DimDepartment ─────────────────────────────────────────────────────────────
def build_dim_department():
    print("Building DimDepartment...")
    rows = [{"department_key":d[0],"department_name":d[1],
              "cost_centre_code":d[2],"division":d[3],
              "region":d[4],"manager":d[5]} for d in DEPARTMENTS]
    print(f"  DimDepartment: {len(rows)} rows")
    return pd.DataFrame(rows)


# ── DimEconomicIndicators ─────────────────────────────────────────────────────
def build_dim_economic_indicators():
    print("Building DimEconomicIndicators...")
    rows = []
    for i, (country, gdp, inf, rate, unem, usd, vix) in enumerate(COUNTRIES):
        for yr in range(2019, 2025):
            rows.append({
                "indicator_key":  i * 6 + (yr - 2019) + 1,
                "date_key":       int(f"{yr}1231"),
                "country":        country,
                "gdp_growth":     rnd(gdp  + gauss(0, 0.5)),
                "inflation_rate": rnd(inf  + gauss(0, 0.4)),
                "policy_rate":    rnd(rate + gauss(0, 0.25)),
                "unemployment_rate": rnd(unem + gauss(0, 0.3)),
                "usd_index":      rnd((usd or 100) + gauss(0, 3), 1) if usd else None,
                "vix":            rnd((vix or 18)  + gauss(0, 4), 1) if vix else None
            })
    print(f"  DimEconomicIndicators: {len(rows)} rows")
    return pd.DataFrame(rows)


# ── FactFinancials ────────────────────────────────────────────────────────────
def build_fact_financials():
    print("Building FactFinancials...")
    base_rev = {1:1567,2:1440,3:1330,4:870,5:970,6:750,7:282,8:376,9:1490,10:2116}
    growth_r = {1:.009,2:.010,3:.011,4:.007,5:.006,6:.008,7:.014,8:.016,9:.005,10:.004}
    rows = []
    for co_key in range(1, 11):
        d = date(2019, 1, 1)
        rev = base_rev[co_key]
        while d <= date(2024, 12, 31):
            rev = rev * (1 + growth_r[co_key] + gauss(0, 0.003))
            cogs = rev * (0.43 + gauss(0, 0.02))
            gp   = rev - cogs
            sga  = rev * (0.14 + gauss(0, 0.01))
            rd   = rev * (0.07 + gauss(0, 0.005))
            ebitda = gp - sga - rd
            da   = rev * 0.04
            ebit = ebitda - da
            interest = rev * 0.02
            ni = (ebit - interest) * 0.79
            rows.append({
                "company_key":     co_key,
                "date_key":        int(d.strftime("%Y%m%d")),
                "fiscal_year":     d.year,
                "revenue":         rnd(rev),
                "cogs":            rnd(cogs),
                "gross_profit":    rnd(gp),
                "sga":             rnd(sga),
                "rd_expense":      rnd(rd),
                "ebitda":          rnd(ebitda),
                "da":              rnd(da),
                "ebit":            rnd(ebit),
                "interest_expense":rnd(interest),
                "net_income":      rnd(ni)
            })
            # Advance to next month
            if d.month == 12:
                d = date(d.year + 1, 1, 1)
            else:
                d = date(d.year, d.month + 1, 1)
    print(f"  FactFinancials: {len(rows):,} rows")
    return pd.DataFrame(rows)


# ── FactBalanceSheet ──────────────────────────────────────────────────────────
def build_fact_balance_sheet():
    print("Building FactBalanceSheet...")
    base_assets = {1:9850,2:9050,3:8400,4:8000,5:5800,
                   6:4600,7:1400,8:1800,9:7400,10:10400}
    rows = []
    for co_key in range(1, 11):
        d = date(2019, 1, 1)
        ta = base_assets[co_key]
        while d <= date(2024, 12, 31):
            ta  = ta * (1 + gauss(0.005, 0.003))
            ca  = ta * (0.38 + gauss(0, 0.02))
            cl  = ca * (0.54 + gauss(0, 0.03))
            debt= ta * (0.20 + gauss(0, 0.02))
            cash= ca * (0.31 + gauss(0, 0.02))
            inv = ca * (0.19 + gauss(0, 0.01))
            rec = ca * (0.25 + gauss(0, 0.01))
            pay = cl * 0.40
            ppe = ta * 0.23
            intang = ta * 0.12
            equity = ta - debt - cl * 0.5
            ret = equity * 0.77
            rows.append({
                "company_key":          co_key,
                "date_key":             int(d.strftime("%Y%m%d")),
                "current_assets":       rnd(ca),
                "current_liabilities":  rnd(cl),
                "total_assets":         rnd(ta),
                "total_debt":           rnd(debt),
                "total_equity":         rnd(equity),
                "cash":                 rnd(cash),
                "inventory":            rnd(inv),
                "receivables":          rnd(rec),
                "payables":             rnd(pay),
                "ppe_net":              rnd(ppe),
                "intangibles":          rnd(intang),
                "retained_earnings":    rnd(ret)
            })
            if d.month == 12:
                d = date(d.year + 1, 1, 1)
            else:
                d = date(d.year, d.month + 1, 1)
    print(f"  FactBalanceSheet: {len(rows):,} rows")
    return pd.DataFrame(rows)


# ── FactMarketPrices ──────────────────────────────────────────────────────────
def build_fact_market_prices():
    print("Building FactMarketPrices...")
    base_price = {1:135,2:220,3:88,4:101,5:52,6:155,7:28,8:38,9:17,10:26}
    shares     = {1:15441943070,2:7432000000,3:5500000000,4:2900000000,
                  5:4270000000,6:2683731854,7:40065428272,8:3771905600,
                  9:3859015385,10:6664338036}
    rows = []
    for tk in range(1, 11):
        p = base_price[tk]
        d = date(2023, 1, 1)
        while d <= date(2024, 12, 31):
            if d.weekday() < 5:
                ret   = gauss(0.0003, 0.018)
                p_op  = p * (1 + gauss(0, 0.005))
                p_cl  = p_op * (1 + ret)
                p_hi  = max(p_op, p_cl) * (1 + abs(gauss(0, 0.006)))
                p_lo  = min(p_op, p_cl) * (1 - abs(gauss(0, 0.006)))
                vol   = int(gauss(40e6, 15e6))
                rows.append({
                    "ticker_key":  tk,
                    "date_key":    int(d.strftime("%Y%m%d")),
                    "open_price":  rnd(p_op, 4),
                    "high_price":  rnd(p_hi, 4),
                    "low_price":   rnd(p_lo, 4),
                    "close_price": rnd(p_cl, 4),
                    "adj_close":   rnd(p_cl * 0.998, 4),
                    "volume":      max(vol, 0),
                    "daily_return":rnd(ret, 6),
                    "market_cap":  rnd(p_cl * shares[tk] / 1e9, 2)
                })
                p = p_cl
            d += timedelta(days=1)
    print(f"  FactMarketPrices: {len(rows):,} rows")
    return pd.DataFrame(rows)


# ── FactPortfolio ─────────────────────────────────────────────────────────────
def build_fact_portfolio():
    print("Building FactPortfolio...")
    weights = {1:.18,2:.14,3:.12,4:.10,5:.08,6:.09,7:.06,8:.07,9:.09,10:.07}
    betas   = {1:1.18,2:.95,3:1.08,4:1.05,5:.88,6:.72,7:.85,8:.92,9:.78,10:.82}
    base_mv = {1:2400,2:1870,3:1600,4:1335,5:1070,6:1200,7:800,8:935,9:1200,10:935}
    rows = []
    d = date(2023, 1, 1)
    while d <= date(2024, 12, 31):
        if d.weekday() < 5:
            bench = gauss(0.00035, 0.012)
            for tk in range(1, 11):
                ret = betas[tk] * bench + gauss(0.0002, 0.008)
                rows.append({
                    "portfolio_key":   1,
                    "ticker_key":      tk,
                    "date_key":        int(d.strftime("%Y%m%d")),
                    "weight":          weights[tk],
                    "daily_return":    rnd(ret, 6),
                    "benchmark_return":rnd(bench, 6),
                    "beta":            betas[tk],
                    "std_dev":         rnd(abs(gauss(0.015, 0.003)), 4),
                    "risk_free_rate":  rnd(0.045 / 252, 8),
                    "market_value":    rnd(base_mv[tk] * (1 + gauss(0.001, 0.01)), 2)
                })
        d += timedelta(days=1)
    print(f"  FactPortfolio: {len(rows):,} rows")
    return pd.DataFrame(rows)


# ── FactCreditRisk ────────────────────────────────────────────────────────────
def build_fact_credit_risk():
    print("Building FactCreditRisk...")
    ratings = ["AAA","AA","A","BBB","BB","B","CCC"]
    sectors = ["Corporate","Retail","SME","Government","Real Estate","Infrastructure"]
    tenors  = ["<1yr","1-3yr","3-5yr",">5yr"]
    pd_map  = {"AAA":.0003,"AA":.0008,"A":.002,"BBB":.0075,"BB":.025,"B":.07,"CCC":.2}
    rows = []
    for i in range(1, 201):
        rating = random.choice(ratings)
        for yr in [2022, 2023, 2024]:
            rows.append({
                "borrower_key":    i,
                "date_key":        int(f"{yr}1231"),
                "rating":          rating,
                "sector":          random.choice(sectors),
                "tenor_band":      random.choice(tenors),
                "ead":             rnd(random.uniform(5, 500)),
                "pd":              rnd(pd_map[rating] * random.uniform(.8, 1.3), 4),
                "lgd":             rnd(random.uniform(.30, .65), 4),
                "collateral_value":rnd(random.uniform(50, 600))
            })
    print(f"  FactCreditRisk: {len(rows):,} rows")
    return pd.DataFrame(rows)


# ── FactFPA ───────────────────────────────────────────────────────────────────
def build_fact_fpa():
    print("Building FactFPA...")
    rows = []
    for dept in DEPARTMENTS:
        d = date(2022, 1, 1)
        while d <= date(2024, 12, 31):
            budget = rnd(random.uniform(80, 1200))
            rows.append({
                "department_key": dept[0],
                "date_key":       int(d.strftime("%Y%m%d")),
                "cost_centre":    dept[1],
                "gl_account":     dept[2],
                "budget":         budget,
                "actual":         rnd(budget * (1 + gauss(.02, .07))),
                "forecast":       rnd(budget * (1 + gauss(.01, .04)))
            })
            if d.month == 12:
                d = date(d.year + 1, 1, 1)
            else:
                d = date(d.year, d.month + 1, 1)
    print(f"  FactFPA: {len(rows):,} rows")
    return pd.DataFrame(rows)


# ── FactTreasury ──────────────────────────────────────────────────────────────
def build_fact_treasury():
    print("Building FactTreasury...")
    rows = []
    for acct in range(1, 6):
        for curr in CURRENCIES:
            d = date(2022, 1, 1)
            cash = random.uniform(100, 900)
            while d <= date(2024, 12, 31):
                if d.weekday() < 5:
                    cash = cash * (1 + gauss(.001, .015))
                    hqla = cash * random.uniform(.7, .9)
                    nco  = cash * random.uniform(.45, .65)
                    rsf  = cash * random.uniform(.7, .85)
                    rows.append({
                        "account_key":           acct,
                        "date_key":              int(d.strftime("%Y%m%d")),
                        "currency_code":         curr,
                        "cash_balance":          rnd(cash),
                        "fx_long":               rnd(cash * random.uniform(.6,.8)),
                        "fx_short":              rnd(cash * random.uniform(.15,.3)),
                        "hqla":                  rnd(hqla),
                        "net_cash_outflows_30d": rnd(nco),
                        "rsf":                   rnd(rsf),
                        "asf":                   rnd(rsf * random.uniform(1.1, 1.3))
                    })
                d += timedelta(days=1)
    print(f"  FactTreasury: {len(rows):,} rows")
    return pd.DataFrame(rows)


# ── FactFraud ─────────────────────────────────────────────────────────────────
def build_fact_fraud():
    print("Building FactFraud...")
    flag_reasons = ["Duplicate Payment","Above Threshold","Off-Hours Transaction",
                    "Benford Deviation","Round Number","Unusual Vendor","Sequential Invoice","Normal"]
    rows = []
    for i in range(1, 3001):
        d = date(2023, 1, 1) + timedelta(days=random.randint(0, 729))
        amount = round(random.expovariate(.002), 2)
        fd     = int(str(amount).replace('.','').lstrip('0')[0])
        is_dup = random.random() < .02
        hour   = random.randint(0, 23)
        score  = random.uniform(1, 10)
        if is_dup: score = min(score + 3, 10)
        if hour < 6 or hour > 22: score = min(score + 1.5, 10)
        rows.append({
            "transaction_key": i,
            "date_key":        int(d.strftime("%Y%m%d")),
            "vendor_key":      random.randint(1, 100),
            "amount":          amount,
            "first_digit":     fd,
            "is_duplicate":    is_dup,
            "is_weekend":      d.weekday() >= 5,
            "hour_of_day":     hour,
            "fraud_score":     rnd(score),
            "flag_reason":     flag_reasons[min(int(score//1.5), 6)] if score > 5 else "Normal"
        })
    print(f"  FactFraud: {len(rows):,} rows")
    return pd.DataFrame(rows)


# ── FactESG ───────────────────────────────────────────────────────────────────
def build_fact_esg():
    print("Building FactESG...")
    base = {1:(82,74,76),2:(78,80,82),3:(71,72,79),4:(64,70,75),5:(42,55,60),
            6:(71,76,78),7:(65,62,68),8:(60,65,70),9:(69,68,75),10:(52,58,62)}
    rows = []
    for co_key in range(1, 11):
        e0, s0, g0 = base[co_key]
        for yr in range(2019, 2025):
            e = min(e0 + gauss(.8,1) * (yr-2018), 100)
            s = min(s0 + gauss(.5,.8) * (yr-2018), 100)
            g = min(g0 + gauss(.4,.6) * (yr-2018), 100)
            rows.append({
                "company_key":         co_key,
                "period_key":          yr * 100 + 12,
                "e_score":             rnd(e, 1),
                "s_score":             rnd(s, 1),
                "g_score":             rnd(g, 1),
                "carbon_intensity":    rnd(max(70-(yr-2018)*3.5+gauss(0,2), 10), 1),
                "renewable_pct":       rnd(min(.15+(yr-2018)*.04+random.uniform(0,.03),1),3),
                "board_diversity_pct": rnd(min(.25+(yr-2018)*.025+random.uniform(0,.02),.6),3),
                "water_usage":         round(random.uniform(400,2000)),
                "employee_turnover":   rnd(random.uniform(.08,.22),3)
            })
    print(f"  FactESG: {len(rows):,} rows")
    return pd.DataFrame(rows)


# ── WRITE TO EXCEL ────────────────────────────────────────────────────────────
def write_to_excel(tables: dict, output_path: str):
    print(f"\nWriting to {output_path}...")
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, df in tables.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            # Style header row
            for col_idx, col in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font      = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
                cell.fill      = PatternFill("solid", fgColor="185FA5")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = Border(
                    bottom=Side(style="thin", color="CCCCCC"),
                    right =Side(style="thin", color="CCCCCC")
                )

            # Auto-fit column widths
            for col in ws.columns:
                max_len   = max((len(str(c.value or "")) for c in col), default=10)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

            ws.row_dimensions[1].height = 28
            ws.freeze_panes = "A2"

    size_mb = os.path.getsize(output_path) / 1024 / 1024
    print(f"Written: {output_path}  ({size_mb:.2f} MB)")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("Enterprise Financial Portfolio — ETL Pipeline")
    print("=" * 60)

    tables = {
        "DimDate":                build_dim_date(),
        "DimCompany":             build_dim_company(),
        "DimTicker":              build_dim_ticker(),
        "DimDepartment":          build_dim_department(),
        "DimEconomicIndicators":  build_dim_economic_indicators(),
        "FactFinancials":         build_fact_financials(),
        "FactBalanceSheet":       build_fact_balance_sheet(),
        "FactMarketPrices":       build_fact_market_prices(),
        "FactPortfolio":          build_fact_portfolio(),
        "FactCreditRisk":         build_fact_credit_risk(),
        "FactFPA":                build_fact_fpa(),
        "FactTreasury":           build_fact_treasury(),
        "FactFraud":              build_fact_fraud(),
        "FactESG":                build_fact_esg(),
    }

    write_to_excel(tables, OUTPUT_PATH)

    print("\n" + "=" * 60)
    print("ETL COMPLETE")
    print("=" * 60)
    total_rows = sum(len(df) for df in tables.values())
    print(f"Total tables: {len(tables)}")
    print(f"Total rows:   {total_rows:,}")
    for name, df in tables.items():
        print(f"  {name:35s} {len(df):>8,} rows")


if __name__ == "__main__":
    main()
